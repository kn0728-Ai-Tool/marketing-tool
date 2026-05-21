# csv_analyzer.py
# =====================================
# CSV分析モジュール
# どんな列構成のCSVでも自動判定し
# AIがトレンド・改善点を抽出する
# =====================================

import json
import io
import re
import pandas as pd
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel


# =====================================
# マーケティング指標の列名候補
# 日本語・英語どちらでも対応する
# =====================================
COLUMN_PATTERNS = {
    "keyword": [
        "キーワード", "keyword", "検索語句", "search term",
        "search query", "クエリ", "query",
    ],
    "clicks": [
        "クリック数", "clicks", "クリック", "click",
    ],
    "impressions": [
        "インプレッション", "impressions", "表示回数", "impr",
    ],
    "ctr": [
        "ctr", "クリック率", "click through rate",
    ],
    "cpc": [
        "平均cpc", "cpc", "avg cpc", "平均クリック単価", "クリック単価",
    ],
    "cost": [
        "費用", "cost", "コスト", "spend", "消化金額",
    ],
    "conversions": [
        "コンバージョン", "conversions", "cv", "成約数", "購入数",
    ],
    "cvr": [
        "cvr", "コンバージョン率", "conversion rate", "転換率",
    ],
    "cpa": [
        "cpa", "コンバージョン単価", "cost per conversion",
    ],
    "roas": [
        "roas", "広告費用対効果",
    ],
    "revenue": [
        "収益", "revenue", "売上", "売上高",
    ],
    "quality_score": [
        "品質スコア", "quality score", "qs",
    ],
    "position": [
        "掲載順位", "position", "avg position", "平均掲載順位",
    ],
}


def detect_columns(df: pd.DataFrame) -> dict:
    """
    DataFrameの列名を走査し、
    マーケティング指標に対応する列名を自動検出して辞書で返す。
    例: {"keyword": "キーワード", "clicks": "クリック数", ...}
    """
    detected = {}
    cols_lower = {c: c.lower().strip() for c in df.columns}

    for metric, candidates in COLUMN_PATTERNS.items():
        for col, col_lower in cols_lower.items():
            if any(c.lower() in col_lower or col_lower in c.lower() for c in candidates):
                detected[metric] = col
                break

    return detected


def clean_numeric(series: pd.Series) -> pd.Series:
    """
    数値列の前処理。
    「%」「¥」「,」などを除去して float に変換する。
    """
    return (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("¥", "", regex=False)
        .str.replace("%", "", regex=False)
        .str.replace(" ", "", regex=False)
        .pipe(pd.to_numeric, errors="coerce")
    )


def prepare_dataframe(df: pd.DataFrame, col_map: dict) -> pd.DataFrame:
    """
    検出した列だけを抽出し、数値列をクリーニングして返す。
    """
    numeric_metrics = [
        "clicks", "impressions", "ctr", "cpc", "cost",
        "conversions", "cvr", "cpa", "roas", "revenue",
        "quality_score", "position",
    ]

    result = pd.DataFrame()

    # キーワード列
    if "keyword" in col_map:
        result["キーワード"] = df[col_map["keyword"]].astype(str).str.strip()

    # 数値列
    for metric in numeric_metrics:
        if metric in col_map:
            label = col_map[metric]
            result[label] = clean_numeric(df[col_map[metric]])

    return result


def build_summary_for_ai(df: pd.DataFrame, col_map: dict, max_rows: int = 30) -> str:
    """
    AIに渡すためのCSVサマリーテキストを生成する。
    行数が多い場合は上位・下位各15件に絞る。
    """
    lines = []

    # 全体統計
    lines.append("## CSVデータの概要")
    lines.append(f"- 総行数: {len(df)}行")
    lines.append(f"- 列構成: {', '.join(df.columns.tolist())}")
    lines.append("")

    # 数値列の基本統計
    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    if numeric_cols:
        lines.append("## 数値列の基本統計")
        for col in numeric_cols:
            s = df[col].dropna()
            if len(s) > 0:
                lines.append(
                    f"- {col}: 平均={s.mean():.2f}, 最大={s.max():.2f}, "
                    f"最小={s.min():.2f}, 合計={s.sum():.2f}"
                )
        lines.append("")

    # データサンプル（上位・下位）
    lines.append("## データサンプル")
    if len(df) > max_rows:
        sample = pd.concat([df.head(max_rows // 2), df.tail(max_rows // 2)])
        lines.append(f"（全{len(df)}行から上位・下位各{max_rows // 2}行を抜粋）")
    else:
        sample = df

    lines.append(sample.to_string(index=False, max_cols=10))

    return "\n".join(lines)


def analyze_csv_with_ai(
    client: OpenAI,
    df: pd.DataFrame,
    col_map: dict,
    industry: str = "",
    custom_question: str = "",
) -> dict:
    """
    CSVデータをAIに渡してトレンド・改善点を分析させる。
    戻り値: {
        "summary":       全体サマリー（文章）,
        "trends":        トレンド分析（リスト）,
        "issues":        課題・問題点（リスト）,
        "improvements":  改善提案（リスト）,
        "top_keywords":  注目キーワード（リスト）,
        "targeting":     ターゲティング提案（リスト）,
        "next_actions":  次のアクション（リスト）,
    }
    """
    data_summary = build_summary_for_ai(df, col_map)
    industry_ctx = f"業種・ジャンル: {industry}\n" if industry else ""
    custom_ctx   = f"特に知りたいこと: {custom_question}\n" if custom_question else ""

    prompt = f"""
あなたはGoogle広告・デジタルマーケティングの第一人者です。
以下のCSVデータを分析し、マーケティング改善に役立つ洞察を提供してください。

{industry_ctx}{custom_ctx}
{data_summary}

以下の形式でJSONのみ出力してください。前置き・説明・```は不要です。

{{
  "summary": "このデータ全体の状況を3〜4文で簡潔にまとめた総評",

  "trends": [
    "トレンド・傾向を具体的な数値を交えて記述（例: CTRが高いキーワード群はXXXの傾向がある）",
    "トレンド2",
    "トレンド3"
  ],

  "issues": [
    "課題・問題点を具体的に記述（例: CPAがXX円を超えているキーワードが○件ある）",
    "課題2",
    "課題3"
  ],

  "improvements": [
    "改善提案を具体的なアクションとして記述（例: CTRが低いXXXキーワードは広告文を○○に変更すべき）",
    "改善提案2",
    "改善提案3",
    "改善提案4",
    "改善提案5"
  ],

  "top_keywords": [
    {{
      "keyword": "注目すべきキーワード名",
      "reason":  "注目する理由（数値根拠を含む）",
      "action":  "このキーワードに対する推奨アクション"
    }}
  ],

  "targeting": [
    "ターゲティング最適化の提案（例: 夜間帯の入札を強化すべき層はXXX）",
    "ターゲティング提案2",
    "ターゲティング提案3"
  ],

  "next_actions": [
    "今すぐやるべきアクション（優先度順）",
    "次のアクション2",
    "次のアクション3",
    "次のアクション4"
  ]
}}
"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "system",
                "content": (
                    "あなたはGoogle広告・デジタルマーケティングの専門家です。"
                    "データに基づいた具体的で実践的な分析・改善提案を行ってください。"
                    "出力はJSONのみ。前置きや説明は一切不要です。"
                ),
            },
            {"role": "user", "content": prompt},
        ],
        temperature=0.7,
        max_tokens=2000,
    )

    raw = response.choices[0].message.content.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


def get_top_bottom_keywords(
    df: pd.DataFrame,
    col_map: dict,
    metric: str = "ctr",
    top_n: int = 5,
) -> tuple:
    """
    指定した指標でキーワードを上位・下位に分けて返す。
    戻り値: (top_df, bottom_df)
    """
    if metric not in col_map or "keyword" not in col_map:
        return pd.DataFrame(), pd.DataFrame()

    col   = col_map[metric]
    kw    = col_map["keyword"]
    valid = df[[kw, col]].dropna().copy()
    valid[col] = clean_numeric(valid[col])
    valid = valid.sort_values(col, ascending=False)

    return valid.head(top_n), valid.tail(top_n)


# =====================================
# Excel出力モジュール
# =====================================
# ここから下が今回追加した新機能です。
# 既存の関数はすべてそのまま維持しています。
# =====================================

# ── スタイル定数 ──────────────────────────
_COLOR_HEADER_BG   = "1E3A5F"   # ヘッダー背景（濃紺）
_COLOR_HEADER_FG   = "FFFFFF"   # ヘッダー文字（白）
_COLOR_SUBHDR_BG   = "2E6DA4"   # サブヘッダー背景（青）
_COLOR_SUBHDR_FG   = "FFFFFF"   # サブヘッダー文字
_COLOR_TITLE_BG    = "1A1A2E"   # タイトル背景（ダーク）
_COLOR_EVEN_ROW    = "EBF5FB"   # 偶数行（薄青）
_COLOR_ODD_ROW     = "FFFFFF"   # 奇数行（白）
_COLOR_SUMMARY_BG  = "FFF3CD"   # サマリーセル背景（薄黄）
_COLOR_ALERT_BG    = "FDECEA"   # 注意値背景（薄赤）
_COLOR_GOOD_BG     = "E8F8F0"   # 良好値背景（薄緑）
_COLOR_ACCENT      = "2980B9"   # アクセント（青）

_FONT_NAME = "Arial"


def _make_thin_border() -> Border:
    """細枠ボーダーを返す"""
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _make_header_border() -> Border:
    """ヘッダー用ボーダー（下線太め）を返す"""
    thick = Side(style="medium", color="FFFFFF")
    thin  = Side(style="thin",   color="AAAAAA")
    return Border(left=thin, right=thin, top=thick, bottom=thick)


def _apply_header_style(cell, text: str, bg: str = _COLOR_HEADER_BG, fg: str = _COLOR_HEADER_FG, size: int = 10):
    """ヘッダーセルにスタイルを適用"""
    cell.value     = text
    cell.font      = Font(name=_FONT_NAME, bold=True, color=fg, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _make_header_border()


def _apply_data_style(cell, value, row_idx: int = 0, fmt: str = None, align: str = "right"):
    """データセルにスタイルを適用"""
    cell.value     = value
    cell.font      = Font(name=_FONT_NAME, size=9)
    bg             = _COLOR_EVEN_ROW if row_idx % 2 == 0 else _COLOR_ODD_ROW
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _make_thin_border()
    if fmt:
        cell.number_format = fmt


def _set_col_widths(ws, widths: dict):
    """列幅を一括設定。widths = {"A": 20, "B": 15, ...}"""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _write_section_title(ws, row: int, col_start: int, col_end: int, title: str):
    """セクションタイトル行を書き込む（結合セル）"""
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start)
    _apply_header_style(cell, title, bg=_COLOR_SUBHDR_BG, size=10)
    ws.row_dimensions[row].height = 22


def _auto_detect_column_type(series: pd.Series) -> str:
    """
    列の値から適切な数値フォーマットを推定する。
    戻り値: "percent" | "yen" | "decimal" | "integer" | "text"
    """
    name_lower = str(series.name).lower()

    # 列名から推定
    if any(k in name_lower for k in ["率", "ctr", "cvr", "roas", "%", "rate"]):
        return "percent"
    if any(k in name_lower for k in ["費用", "cpc", "cpa", "cost", "spend", "売上", "revenue", "単価", "¥"]):
        return "yen"

    # 値から推定
    s = series.dropna()
    if len(s) == 0:
        return "text"
    try:
        nums = pd.to_numeric(s, errors="coerce").dropna()
        if len(nums) == 0:
            return "text"
        if nums.between(0, 1).mean() > 0.8:
            return "percent"
        if nums.max() > 100:
            return "integer"
        return "decimal"
    except Exception:
        return "text"


def _get_number_format(col_type: str) -> str:
    """列タイプに対応するExcel数値フォーマット文字列を返す"""
    return {
        "percent": "0.00%",
        "yen":     "#,##0",
        "decimal": "0.00",
        "integer": "#,##0",
        "text":    "@",
    }.get(col_type, "General")


# ── データクリーニング ─────────────────────

def clean_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    DataFrameを徹底クリーニングして返す。

    処理内容：
    1. 完全に空の行・列を削除
    2. 列名の前後スペースを除去
    3. 重複行を削除
    4. 文字列セルの前後スペース・改行を除去
    5. 数値列の表記ゆれ（¥, %, カンマ）を統一

    戻り値: (cleaned_df, report_dict)
    report_dict = {
        "removed_empty_rows":   int,
        "removed_empty_cols":   int,
        "removed_dup_rows":     int,
        "cleaned_str_cells":    int,
        "cleaned_num_cells":    int,
        "original_shape":       (int, int),
        "cleaned_shape":        (int, int),
    }
    """
    report = {
        "removed_empty_rows":  0,
        "removed_empty_cols":  0,
        "removed_dup_rows":    0,
        "cleaned_str_cells":   0,
        "cleaned_num_cells":   0,
        "original_shape":      df.shape,
        "cleaned_shape":       (0, 0),
    }

    work = df.copy()

    # 1. 列名クリーニング
    work.columns = [str(c).strip() for c in work.columns]

    # 2. 完全に空の列を削除
    before_cols = len(work.columns)
    work = work.dropna(axis=1, how="all")
    # 全セルが空文字 or スペースの列も削除
    for col in work.columns.tolist():
        if work[col].astype(str).str.strip().eq("").all():
            work = work.drop(columns=[col])
    report["removed_empty_cols"] = before_cols - len(work.columns)

    # 3. 完全に空の行を削除
    before_rows = len(work)
    work = work.dropna(how="all")
    # 全セルが空文字 or スペースの行も削除
    mask = work.apply(lambda row: row.astype(str).str.strip().eq("").all(), axis=1)
    work = work[~mask]
    report["removed_empty_rows"] = before_rows - len(work)

    # 4. 重複行を削除
    before_dup = len(work)
    work = work.drop_duplicates()
    report["removed_dup_rows"] = before_dup - len(work)

    # 5. 文字列列のクリーニング（前後スペース・改行除去）
    str_cleaned = 0
    for col in work.columns:
        if work[col].dtype == object:
            original = work[col].astype(str)
            cleaned  = original.str.strip().str.replace(r"\s+", " ", regex=True)
            diff = (original != cleaned).sum()
            str_cleaned += int(diff)
            work[col] = cleaned
    report["cleaned_str_cells"] = str_cleaned

    # 6. 数値列の表記ゆれ統一（¥ % カンマ 全角数字）
    num_cleaned = 0
    _fullwidth = str.maketrans("０１２３４５６７８９．", "0123456789.")
    for col in work.columns:
        sample = work[col].astype(str).str.strip()
        # 数値っぽい列を検出（¥や%や,を除いた状態で数値になるか確認）
        test = (
            sample
            .str.replace(",", "", regex=False)
            .str.replace("¥", "", regex=False)
            .str.replace("￥", "", regex=False)
            .str.replace("%", "", regex=False)
            .str.replace(" ", "", regex=False)
            .str.translate(_fullwidth)
        )
        numeric_ratio = pd.to_numeric(test, errors="coerce").notna().mean()
        if numeric_ratio > 0.5:
            converted = pd.to_numeric(test, errors="coerce")
            non_null_orig = work[col].notna().sum()
            non_null_new  = converted.notna().sum()
            if non_null_new >= non_null_orig * 0.8:
                work[col] = converted
                num_cleaned += int(non_null_new)
    report["cleaned_num_cells"] = num_cleaned

    # インデックスをリセット
    work = work.reset_index(drop=True)
    report["cleaned_shape"] = work.shape

    return work, report


# ── シート生成ヘルパー ─────────────────────

def _write_clean_report_sheet(wb: Workbook, report: dict):
    """「クリーニングレポート」シートを生成"""
    ws = wb.create_sheet("📋 クリーニングレポート")
    ws.sheet_view.showGridLines = False

    # タイトル
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = "データクリーニングレポート"
    c.font      = Font(name=_FONT_NAME, bold=True, color=_COLOR_HEADER_FG, size=13)
    c.fill      = PatternFill("solid", fgColor=_COLOR_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    items = [
        ("元データ（行×列）", f"{report['original_shape'][0]}行 × {report['original_shape'][1]}列"),
        ("クリーニング後",     f"{report['cleaned_shape'][0]}行 × {report['cleaned_shape'][1]}列"),
        ("削除した空白行",     f"{report['removed_empty_rows']}行"),
        ("削除した空白列",     f"{report['removed_empty_cols']}列"),
        ("削除した重複行",     f"{report['removed_dup_rows']}行"),
        ("整形した文字列セル", f"{report['cleaned_str_cells']}セル"),
        ("数値に変換したセル", f"{report['cleaned_num_cells']}セル"),
    ]

    for i, (label, val) in enumerate(items, start=3):
        row_idx = i - 3
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=val)
        for cell in (lc, vc):
            cell.font      = Font(name=_FONT_NAME, size=10)
            cell.fill      = PatternFill("solid", fgColor=_COLOR_EVEN_ROW if row_idx % 2 == 0 else _COLOR_ODD_ROW)
            cell.border    = _make_thin_border()
            cell.alignment = Alignment(horizontal="left", vertical="center")
        lc.font = Font(name=_FONT_NAME, bold=True, size=10)
        ws.row_dimensions[i].height = 20

    _set_col_widths(ws, {"A": 28, "B": 30})


def _write_summary_sheet(wb: Workbook, df: pd.DataFrame, file_name: str = ""):
    """「サマリー」シートを生成（全体の基本統計）"""
    ws = wb.create_sheet("📊 サマリー")
    ws.sheet_view.showGridLines = False

    # タイトル
    num_cols = max(len(df.columns) + 1, 5)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws["A1"]
    c.value     = f"集計サマリー｜{file_name}　（全{len(df)}行 × {len(df.columns)}列）"
    c.font      = Font(name=_FONT_NAME, bold=True, color=_COLOR_HEADER_FG, size=12)
    c.fill      = PatternFill("solid", fgColor=_COLOR_TITLE_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # 数値列の統計
    numeric_cols = df.select_dtypes(include="number").columns.tolist()

    if numeric_cols:
        _write_section_title(ws, 3, 1, 7, "▼ 数値列の基本統計")
        headers = ["列名", "件数", "合計", "平均", "最大", "最小", "標準偏差"]
        for ci, h in enumerate(headers, start=1):
            _apply_header_style(ws.cell(row=4, column=ci), h, bg=_COLOR_SUBHDR_BG)
        ws.row_dimensions[4].height = 20

        for ri, col in enumerate(numeric_cols):
            s = df[col].dropna()
            col_type = _auto_detect_column_type(s.rename(col))
            fmt      = _get_number_format(col_type)
            row      = 5 + ri
            vals = [col, len(s), s.sum(), s.mean(), s.max(), s.min(), s.std()]
            fmts = ["@", "#,##0", fmt, fmt, fmt, fmt, fmt]
            for ci, (v, f) in enumerate(zip(vals, fmts), start=1):
                _apply_data_style(ws.cell(row=row, column=ci), v, row_idx=ri, fmt=f,
                                  align="left" if ci == 1 else "right")
            ws.row_dimensions[row].height = 18

        # 数値列のEXCEL数式で合計行を追加
        total_row = 5 + len(numeric_cols)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=1)
        c = ws.cell(row=total_row, column=1, value="※ 上記はクリーニング後のデータに基づく集計です")
        c.font      = Font(name=_FONT_NAME, italic=True, color="888888", size=8)
        c.alignment = Alignment(horizontal="left")

    # テキスト列の情報
    text_cols = df.select_dtypes(exclude="number").columns.tolist()
    text_start = 5 + len(numeric_cols) + 3

    if text_cols:
        _write_section_title(ws, text_start, 1, 5, "▼ テキスト列の情報")
        t_headers = ["列名", "ユニーク件数", "空白件数", "最頻値", "最頻値の出現回数"]
        for ci, h in enumerate(t_headers, start=1):
            _apply_header_style(ws.cell(row=text_start + 1, column=ci), h, bg=_COLOR_SUBHDR_BG)
        ws.row_dimensions[text_start + 1].height = 20

        for ri, col in enumerate(text_cols):
            s = df[col].astype(str).replace("nan", "")
            unique_count  = s[s != ""].nunique()
            empty_count   = (s.str.strip() == "").sum()
            vc = s[s != ""].value_counts()
            top_val   = vc.index[0]   if len(vc) > 0 else "-"
            top_count = int(vc.iloc[0]) if len(vc) > 0 else 0
            row = text_start + 2 + ri
            vals = [col, unique_count, empty_count, top_val, top_count]
            fmts = ["@", "#,##0", "#,##0", "@", "#,##0"]
            aligns = ["left", "right", "right", "left", "right"]
            for ci, (v, f, a) in enumerate(zip(vals, fmts, aligns), start=1):
                _apply_data_style(ws.cell(row=row, column=ci), v, row_idx=ri, fmt=f, align=a)
            ws.row_dimensions[row].height = 18

    _set_col_widths(ws, {"A": 25, "B": 12, "C": 16, "D": 14, "E": 14, "F": 14, "G": 14})


def _write_data_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str):
    """
    元データ（クリーニング済み）シートを生成。
    全列をヘッダー付きで出力し、交互行色でテーブルを整える。
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"  # ヘッダー行を固定

    # ヘッダー行
    for ci, col in enumerate(df.columns, start=1):
        _apply_header_style(ws.cell(row=1, column=ci), col)
    ws.row_dimensions[1].height = 22

    # データ行
    for ri, (_, row_data) in enumerate(df.iterrows()):
        for ci, col in enumerate(df.columns, start=1):
            val = row_data[col]
            # NaN→空文字
            if pd.isna(val):
                val = ""
            col_type = _auto_detect_column_type(df[col])
            fmt = _get_number_format(col_type) if col_type != "text" else "@"
            align = "left" if col_type == "text" else "right"
            _apply_data_style(ws.cell(row=ri + 2, column=ci), val, row_idx=ri, fmt=fmt, align=align)
        ws.row_dimensions[ri + 2].height = 16

    # 列幅の自動調整（最大30文字）
    for ci, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 30)

    # オートフィルター
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"


def _write_column_detail_sheet(wb: Workbook, df: pd.DataFrame, col_name: str, col_type: str):
    """
    1列分の詳細集計シートを生成する。
    数値列なら統計＋上位/下位ランキング、
    テキスト列なら値別の出現回数を集計する。
    """
    # シート名は最大31文字（Excelの制限）・特殊文字を除去
    safe_name = re.sub(r'[\\/*?:\[\]]', '_', col_name)[:29]

    ws = wb.create_sheet(f"📌 {safe_name}")
    ws.sheet_view.showGridLines = False

    fmt = _get_number_format(col_type)

    # ── タイトル ───────────────────────────
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = f"列詳細：{col_name}"
    c.font      = Font(name=_FONT_NAME, bold=True, color=_COLOR_HEADER_FG, size=12)
    c.fill      = PatternFill("solid", fgColor=_COLOR_TITLE_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    if col_type in ("percent", "yen", "decimal", "integer"):
        # ── 数値列の詳細 ──────────────────────
        s = df[col_name].dropna()

        _write_section_title(ws, 3, 1, 3, "▼ 基本統計")
        stats = [
            ("件数（データあり）", len(s)),
            ("件数（空白）",       len(df) - len(s)),
            ("合計",               f"=SUM(C8:C{8 + len(s) - 1})" if len(s) > 0 else 0),
            ("平均",               f"=AVERAGE(C8:C{8 + len(s) - 1})" if len(s) > 0 else 0),
            ("最大値",             s.max() if len(s) > 0 else 0),
            ("最小値",             s.min() if len(s) > 0 else 0),
            ("中央値",             s.median() if len(s) > 0 else 0),
            ("標準偏差",           s.std() if len(s) > 0 else 0),
        ]
        _apply_header_style(ws.cell(row=4, column=1), "項目",  bg=_COLOR_SUBHDR_BG)
        _apply_header_style(ws.cell(row=4, column=2), "値",    bg=_COLOR_SUBHDR_BG)
        for ri2, (label, val) in enumerate(stats):
            row = 5 + ri2
            lc = ws.cell(row=row, column=1, value=label)
            vc = ws.cell(row=row, column=2, value=val)
            bg = _COLOR_EVEN_ROW if ri2 % 2 == 0 else _COLOR_ODD_ROW
            for cell in (lc, vc):
                cell.font   = Font(name=_FONT_NAME, size=9)
                cell.fill   = PatternFill("solid", fgColor=bg)
                cell.border = _make_thin_border()
            lc.alignment = Alignment(horizontal="left",  vertical="center")
            vc.alignment = Alignment(horizontal="right", vertical="center")
            if isinstance(val, (int, float)):
                vc.number_format = fmt
            ws.row_dimensions[row].height = 18

        # 上位20件テーブル（あいうえお順や数値上位）
        sorted_df = df[["キーワード", col_name]].dropna() if "キーワード" in df.columns else df[[col_name]].dropna()
        sorted_df = sorted_df.sort_values(col_name, ascending=False).head(20).reset_index(drop=True)

        data_start = 8
        label = "▼ 上位20件"
        _write_section_title(ws, data_start - 1, 1, 3, label)
        for ci2, col in enumerate(sorted_df.columns, start=1):
            _apply_header_style(ws.cell(row=data_start, column=ci2), col, bg=_COLOR_SUBHDR_BG)

        for ri2, (_, row_data) in enumerate(sorted_df.iterrows()):
            for ci2, col in enumerate(sorted_df.columns, start=1):
                val = row_data[col]
                if pd.isna(val):
                    val = ""
                cell = ws.cell(row=data_start + 1 + ri2, column=ci2)
                col_t = _auto_detect_column_type(sorted_df[col])
                f = _get_number_format(col_t) if col_t != "text" else "@"
                a = "left" if col_t == "text" else "right"
                _apply_data_style(cell, val, row_idx=ri2, fmt=f, align=a)
                # トップ3を強調
                if ci2 == len(sorted_df.columns) and ri2 < 3:
                    cell.fill = PatternFill("solid", fgColor=_COLOR_GOOD_BG)
                    cell.font = Font(name=_FONT_NAME, bold=True, size=9)
            ws.row_dimensions[data_start + 1 + ri2].height = 16

        # ── シンプルな棒グラフ ──
        if len(sorted_df) >= 2 and col_name in sorted_df.columns:
            chart = BarChart()
            chart.type        = "bar"
            chart.title       = f"{col_name} 上位20件"
            chart.y_axis.title = col_name
            chart.x_axis.title = "キーワード" if "キーワード" in sorted_df.columns else "行"
            chart.width  = 22
            chart.height = 14
            chart.style  = 10
            chart.grouping = "clustered"
            chart.overlap  = 100

            val_col   = len(sorted_df.columns)
            label_col = 1

            data_ref = Reference(
                ws,
                min_col=val_col,
                min_row=data_start,
                max_row=data_start + len(sorted_df),
            )
            cats_ref = Reference(
                ws,
                min_col=label_col,
                min_row=data_start + 1,
                max_row=data_start + len(sorted_df),
            )
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, f"E{data_start}")

    else:
        # ── テキスト列の詳細（値別出現回数） ──────
        s = df[col_name].astype(str).str.strip().replace("nan", "").replace("", pd.NA).dropna()
        vc = s.value_counts().reset_index()
        vc.columns = [col_name, "件数"]

        _write_section_title(ws, 3, 1, 3, "▼ 値別 出現回数（降順）")
        headers = [col_name, "件数", "割合"]
        for ci2, h in enumerate(headers, start=1):
            _apply_header_style(ws.cell(row=4, column=ci2), h, bg=_COLOR_SUBHDR_BG)

        total = vc["件数"].sum()
        for ri2, (_, row_data) in enumerate(vc.iterrows()):
            row = 5 + ri2
            val_text  = row_data[col_name]
            val_count = int(row_data["件数"])
            val_pct   = val_count / total if total > 0 else 0
            vals  = [val_text, val_count, val_pct]
            fmts  = ["@", "#,##0", "0.00%"]
            aligns = ["left", "right", "right"]
            for ci2, (v, f, a) in enumerate(zip(vals, fmts, aligns), start=1):
                _apply_data_style(ws.cell(row=row, column=ci2), v, row_idx=ri2, fmt=f, align=a)
            ws.row_dimensions[row].height = 16

    _set_col_widths(ws, {"A": 28, "B": 16, "C": 12, "D": 12, "E": 4})


# ── メイン関数 ─────────────────────────────

def export_to_excel(
    df_raw: pd.DataFrame,
    file_name: str = "データ",
    ai_result: dict = None,
) -> bytes:
    """
    CSVデータをクリーニングし、項目別シートを持つExcelファイルを生成して
    バイト列で返す。Streamlitのdownload_buttonに直接渡せる。

    シート構成：
      1. 📋 クリーニングレポート  ── 削除/修正した内容のサマリー
      2. 📊 サマリー             ── 全列の基本統計
      3. 🗂️ クリーニング済みデータ ── 元データ（全列・ヘッダー固定）
      4. 📌 {列名} × 列の数      ── 各列の詳細集計（数値:統計+ランキング / テキスト:頻度表）
      5. 💡 AI分析結果            ── ai_resultがある場合のみ

    Parameters
    ----------
    df_raw    : アップロードされたCSVのDataFrame（クリーニング前）
    file_name : ファイル名表示用（シートタイトルなどに使用）
    ai_result : analyze_csv_with_ai() の戻り値（任意）

    Returns
    -------
    bytes : xlsx形式のバイト列
    """
    # ── 1. クリーニング ──
    df_clean, clean_report = clean_dataframe(df_raw)

    # ── 2. ワークブック生成 ──
    wb = Workbook()
    # デフォルトシートを削除
    default_sheet = wb.active
    wb.remove(default_sheet)

    # ── 3. 各シートを生成 ──
    _write_clean_report_sheet(wb, clean_report)
    _write_summary_sheet(wb, df_clean, file_name)
    _write_data_sheet(wb, df_clean, "🗂️ クリーニング済みデータ")

    # 列ごとの詳細シート（最大20列まで。それ以上は主要列のみ）
    max_detail_cols = 20
    cols_to_detail = df_clean.columns.tolist()[:max_detail_cols]

    for col in cols_to_detail:
        col_type = _auto_detect_column_type(df_clean[col])
        _write_column_detail_sheet(wb, df_clean, col, col_type)

    # ── 4. AI分析結果シート（任意） ──
    if ai_result and not ai_result.get("error"):
        _write_ai_result_sheet(wb, ai_result)

    # ── 5. バイト列に変換して返す ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _write_ai_result_sheet(wb: Workbook, ai: dict):
    """AI分析結果シートを生成"""
    ws = wb.create_sheet("💡 AI分析結果")
    ws.sheet_view.showGridLines = False

    # タイトル
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = "AI分析結果"
    c.font      = Font(name=_FONT_NAME, bold=True, color=_COLOR_HEADER_FG, size=13)
    c.fill      = PatternFill("solid", fgColor=_COLOR_TITLE_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    sections = [
        ("📋 総評",             [ai.get("summary", "")]),
        ("📈 トレンド・傾向",    ai.get("trends", [])),
        ("⚠️ 課題・問題点",      ai.get("issues", [])),
        ("🔧 改善提案",          ai.get("improvements", [])),
        ("🎯 ターゲティング提案", ai.get("targeting", [])),
        ("✅ 今すぐやるべきアクション", ai.get("next_actions", [])),
    ]

    current_row = 3
    for section_title, items in sections:
        # セクションヘッダー
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=2
        )
        _apply_header_style(ws.cell(row=current_row, column=1), section_title, bg=_COLOR_SUBHDR_BG)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for i, item in enumerate(items if isinstance(items, list) else [items]):
            if not item:
                continue
            c = ws.cell(row=current_row, column=1, value=f"{'🥇🥈🥉'[i] if i < 3 else '▶'}" if "アクション" in section_title else "·")
            c.font      = Font(name=_FONT_NAME, size=9, bold=True, color=_COLOR_ACCENT)
            c.fill      = PatternFill("solid", fgColor=_COLOR_EVEN_ROW if i % 2 == 0 else _COLOR_ODD_ROW)
            c.alignment = Alignment(horizontal="center", vertical="top")
            c.border    = _make_thin_border()

            tc = ws.cell(row=current_row, column=2, value=str(item))
            tc.font      = Font(name=_FONT_NAME, size=9)
            tc.fill      = PatternFill("solid", fgColor=_COLOR_EVEN_ROW if i % 2 == 0 else _COLOR_ODD_ROW)
            tc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            tc.border    = _make_thin_border()
            ws.row_dimensions[current_row].height = 32
            current_row += 1

        current_row += 1  # セクション間の空白行

    # 注目キーワード
    top_kws = ai.get("top_keywords", [])
    if top_kws:
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=3
        )
        _apply_header_style(ws.cell(row=current_row, column=1), "🔑 注目キーワード", bg=_COLOR_SUBHDR_BG)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for h, col_idx in [("キーワード", 1), ("注目理由", 2), ("推奨アクション", 3)]:
            _apply_header_style(ws.cell(row=current_row, column=col_idx), h, bg=_COLOR_HEADER_BG)
        current_row += 1

        for ri2, kw in enumerate(top_kws):
            vals = [kw.get("keyword", ""), kw.get("reason", ""), kw.get("action", "")]
            for ci2, val in enumerate(vals, start=1):
                tc = ws.cell(row=current_row, column=ci2, value=val)
                tc.font      = Font(name=_FONT_NAME, size=9)
                tc.fill      = PatternFill("solid", fgColor=_COLOR_EVEN_ROW if ri2 % 2 == 0 else _COLOR_ODD_ROW)
                tc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                tc.border    = _make_thin_border()
            ws.row_dimensions[current_row].height = 28
            current_row += 1

    _set_col_widths(ws, {"A": 6, "B": 60, "C": 50})
